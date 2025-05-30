from typing import Dict, List, Optional, Any, Tuple
import json
import csv
from datetime import datetime
import logging
from pathlib import Path
import io

# Optional dependencies for enhanced report generation
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.piecharts import Pie
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False
    logging.warning("reportlab not installed. PDF generation will be unavailable.")

# Optional dependencies for Excel generation
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.chart import LineChart, BarChart, PieChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logging.warning("openpyxl not installed. Excel generation will be unavailable.")

# Optional dependencies for HTML reports
try:
    import pandas as pd
    import plotly.graph_objects as go
    import plotly.express as px
    from plotly.offline import plot
    HAS_PLOTLY = True
except ImportError:
    HAS_PLOTLY = False
    logging.warning("plotly/pandas not installed. Interactive HTML reports will be unavailable.")

logger = logging.getLogger(__name__)


class ReportGenerator:
    """
    Generates comprehensive reports for H2H battles and league analysis.
    Supports multiple output formats: JSON, CSV, PDF.
    """
    
    def __init__(self, live_data_service, h2h_analyzer, enhanced_h2h_analyzer=None):
        """
        Initialize the report generator with required services.
        
        Args:
            live_data_service: Service for fetching FPL data
            h2h_analyzer: Core H2H analysis service
            enhanced_h2h_analyzer: Enhanced analytics service (optional)
        """
        self.live_data_service = live_data_service
        self.h2h_analyzer = h2h_analyzer
        self.enhanced_h2h_analyzer = enhanced_h2h_analyzer
        
        # Create reports directory if it doesn't exist
        self.reports_dir = Path("reports")
        self.reports_dir.mkdir(exist_ok=True)
    
    async def generate_h2h_season_report(
        self,
        manager1_id: int,
        manager2_id: int,
        league_id: int,
        output_format: str = "json"
    ) -> Dict[str, Any]:
        """
        Generate a comprehensive season report for H2H battles between two managers.
        
        Args:
            manager1_id: First manager ID
            manager2_id: Second manager ID
            league_id: H2H league ID
            output_format: Output format (json, csv, pdf)
            
        Returns:
            Dict containing report data and file path
        """
        logger.info(f"Generating H2H season report for managers {manager1_id} vs {manager2_id}")
        
        # Gather all season data
        report_data = await self._gather_season_data(manager1_id, manager2_id, league_id)
        
        # Generate report in requested format
        if output_format.lower() == "json":
            file_path = await self._generate_json_report(report_data, manager1_id, manager2_id)
        elif output_format.lower() == "csv":
            file_path = await self._generate_csv_report(report_data, manager1_id, manager2_id)
        elif output_format.lower() == "pdf":
            if not HAS_REPORTLAB:
                raise ValueError("PDF generation requires reportlab to be installed")
            file_path = await self._generate_enhanced_pdf_report(report_data, manager1_id, manager2_id)
        elif output_format.lower() == "excel" or output_format.lower() == "xlsx":
            if not HAS_OPENPYXL:
                raise ValueError("Excel generation requires openpyxl to be installed")
            file_path = await self._generate_excel_report(report_data, manager1_id, manager2_id)
        elif output_format.lower() == "html":
            if not HAS_PLOTLY:
                raise ValueError("HTML generation requires plotly and pandas to be installed")
            file_path = await self._generate_html_report(report_data, manager1_id, manager2_id)
        else:
            raise ValueError(f"Unsupported output format: {output_format}")
        
        return {
            "status": "success",
            "file_path": str(file_path),
            "format": output_format,
            "data": report_data
        }
    
    async def _gather_season_data(
        self,
        manager1_id: int,
        manager2_id: int,
        league_id: int
    ) -> Dict[str, Any]:
        """Gather all season data for the H2H report."""
        
        # Get manager information
        manager1_info = await self.live_data_service.get_manager_info(manager1_id)
        manager2_info = await self.live_data_service.get_manager_info(manager2_id)
        
        # Get all H2H matches between these managers
        all_matches = []
        h2h_record = {"manager1_wins": 0, "manager2_wins": 0, "draws": 0}
        
        # Fetch matches for all gameweeks in batches for better performance
        import asyncio
        
        async def fetch_gw_matches(gw):
            try:
                matches = await self.h2h_analyzer.get_h2h_matches(league_id, gw)
                return gw, matches
            except Exception as e:
                logger.warning(f"Error fetching matches for GW{gw}: {e}")
                return gw, []
        
        # Create tasks for all gameweeks and run them concurrently
        tasks = [fetch_gw_matches(gw) for gw in range(1, 39)]
        gw_results = await asyncio.gather(*tasks, return_exceptions=True)
        
        # Process results
        for gw_result in gw_results:
            if isinstance(gw_result, Exception):
                continue
                
            gw, matches = gw_result
            for match in matches:
                if self._is_match_between_managers(match, manager1_id, manager2_id):
                    all_matches.append(match)
                    # Update H2H record
                    if match.get('entry_1_entry') == manager1_id:
                        if match.get('entry_1_points', 0) > match.get('entry_2_points', 0):
                            h2h_record["manager1_wins"] += 1
                        elif match.get('entry_1_points', 0) < match.get('entry_2_points', 0):
                            h2h_record["manager2_wins"] += 1
                        else:
                            h2h_record["draws"] += 1
                    else:
                        if match.get('entry_2_points', 0) > match.get('entry_1_points', 0):
                            h2h_record["manager1_wins"] += 1
                        elif match.get('entry_2_points', 0) < match.get('entry_1_points', 0):
                            h2h_record["manager2_wins"] += 1
                        else:
                            h2h_record["draws"] += 1
        
        # Calculate statistics
        stats = self._calculate_season_statistics(all_matches, manager1_id, manager2_id)
        
        # Get key battles (biggest wins, closest matches)
        key_battles = self._identify_key_battles(all_matches, manager1_id, manager2_id)
        
        # Get form analysis
        form_analysis = self._analyze_form(all_matches, manager1_id, manager2_id)
        
        # If enhanced analyzer is available, get advanced analytics
        advanced_analytics = None
        if self.enhanced_h2h_analyzer:
            try:
                # Get analytics for the most recent match
                if all_matches:
                    latest_match = all_matches[-1]
                    latest_gw = latest_match.get('event', 38)
                    advanced_analytics = await self.enhanced_h2h_analyzer.analyze_battle_comprehensive(
                        manager1_id, manager2_id, latest_gw
                    )
            except Exception as e:
                logger.warning(f"Error getting advanced analytics: {e}")
        
        return {
            "managers": {
                "manager1": {
                    "id": manager1_id,
                    "name": manager1_info.get('name', ''),
                    "player_name": f"{manager1_info.get('player_first_name', '')} {manager1_info.get('player_last_name', '')}".strip()
                },
                "manager2": {
                    "id": manager2_id,
                    "name": manager2_info.get('name', ''),
                    "player_name": f"{manager2_info.get('player_first_name', '')} {manager2_info.get('player_last_name', '')}".strip()
                }
            },
            "h2h_record": h2h_record,
            "all_matches": all_matches,
            "statistics": stats,
            "key_battles": key_battles,
            "form_analysis": form_analysis,
            "advanced_analytics": advanced_analytics,
            "generated_at": datetime.utcnow().isoformat()
        }
    
    def _is_match_between_managers(self, match: Dict[str, Any], manager1_id: int, manager2_id: int) -> bool:
        """Check if a match is between the two specified managers."""
        entry1 = match.get('entry_1_entry')
        entry2 = match.get('entry_2_entry')
        return (entry1 == manager1_id and entry2 == manager2_id) or \
               (entry1 == manager2_id and entry2 == manager1_id)
    
    def _calculate_season_statistics(
        self,
        matches: List[Dict[str, Any]],
        manager1_id: int,
        manager2_id: int
    ) -> Dict[str, Any]:
        """Calculate season statistics from all matches."""
        if not matches:
            return {}
        
        manager1_points = []
        manager2_points = []
        margins = []
        
        for match in matches:
            if match.get('entry_1_entry') == manager1_id:
                m1_pts = match.get('entry_1_points', 0)
                m2_pts = match.get('entry_2_points', 0)
            else:
                m1_pts = match.get('entry_2_points', 0)
                m2_pts = match.get('entry_1_points', 0)
            
            manager1_points.append(m1_pts)
            manager2_points.append(m2_pts)
            margins.append(m1_pts - m2_pts)
        
        return {
            "manager1": {
                "total_points": sum(manager1_points),
                "avg_points": sum(manager1_points) / len(manager1_points) if manager1_points else 0,
                "highest_score": max(manager1_points) if manager1_points else 0,
                "lowest_score": min(manager1_points) if manager1_points else 0
            },
            "manager2": {
                "total_points": sum(manager2_points),
                "avg_points": sum(manager2_points) / len(manager2_points) if manager2_points else 0,
                "highest_score": max(manager2_points) if manager2_points else 0,
                "lowest_score": min(manager2_points) if manager2_points else 0
            },
            "margins": {
                "avg_margin": sum(margins) / len(margins) if margins else 0,
                "biggest_win_margin": max(margins) if margins else 0,
                "biggest_loss_margin": min(margins) if margins else 0
            }
        }
    
    def _identify_key_battles(
        self,
        matches: List[Dict[str, Any]],
        manager1_id: int,
        manager2_id: int
    ) -> Dict[str, Any]:
        """Identify key battles from the season."""
        if not matches:
            return {}
        
        # Find biggest wins for each manager
        biggest_m1_win = None
        biggest_m2_win = None
        closest_match = None
        highest_scoring = None
        
        max_m1_margin = -999
        max_m2_margin = -999
        min_margin = 999
        max_total = 0
        
        for match in matches:
            if match.get('entry_1_entry') == manager1_id:
                m1_pts = match.get('entry_1_points', 0)
                m2_pts = match.get('entry_2_points', 0)
            else:
                m1_pts = match.get('entry_2_points', 0)
                m2_pts = match.get('entry_1_points', 0)
            
            margin = m1_pts - m2_pts
            total = m1_pts + m2_pts
            
            # Biggest manager1 win
            if margin > max_m1_margin:
                max_m1_margin = margin
                biggest_m1_win = match
            
            # Biggest manager2 win
            if margin < max_m2_margin:
                max_m2_margin = margin
                biggest_m2_win = match
            
            # Closest match
            if abs(margin) < abs(min_margin):
                min_margin = margin
                closest_match = match
            
            # Highest scoring
            if total > max_total:
                max_total = total
                highest_scoring = match
        
        return {
            "biggest_manager1_win": biggest_m1_win,
            "biggest_manager2_win": biggest_m2_win,
            "closest_match": closest_match,
            "highest_scoring_match": highest_scoring
        }
    
    def _analyze_form(
        self,
        matches: List[Dict[str, Any]],
        manager1_id: int,
        manager2_id: int
    ) -> Dict[str, Any]:
        """Analyze form trends throughout the season."""
        if not matches:
            return {}
        
        # Sort matches by gameweek
        sorted_matches = sorted(matches, key=lambda x: x.get('event', 0))
        
        # Calculate form streaks
        current_m1_streak = 0
        current_m2_streak = 0
        best_m1_streak = 0
        best_m2_streak = 0
        
        for match in sorted_matches:
            if match.get('entry_1_entry') == manager1_id:
                m1_pts = match.get('entry_1_points', 0)
                m2_pts = match.get('entry_2_points', 0)
            else:
                m1_pts = match.get('entry_2_points', 0)
                m2_pts = match.get('entry_1_points', 0)
            
            if m1_pts > m2_pts:
                current_m1_streak += 1
                current_m2_streak = 0
                best_m1_streak = max(best_m1_streak, current_m1_streak)
            elif m2_pts > m1_pts:
                current_m2_streak += 1
                current_m1_streak = 0
                best_m2_streak = max(best_m2_streak, current_m2_streak)
            else:
                # Draw doesn't break streaks
                pass
        
        # Last 5 matches form
        last_5_results = []
        for match in sorted_matches[-5:]:
            if match.get('entry_1_entry') == manager1_id:
                m1_pts = match.get('entry_1_points', 0)
                m2_pts = match.get('entry_2_points', 0)
            else:
                m1_pts = match.get('entry_2_points', 0)
                m2_pts = match.get('entry_1_points', 0)
            
            if m1_pts > m2_pts:
                last_5_results.append('W')
            elif m2_pts > m1_pts:
                last_5_results.append('L')
            else:
                last_5_results.append('D')
        
        return {
            "best_winning_streak": {
                "manager1": best_m1_streak,
                "manager2": best_m2_streak
            },
            "current_form": {
                "manager1": ''.join(last_5_results),
                "manager2": ''.join(['W' if r == 'L' else 'L' if r == 'W' else 'D' for r in last_5_results])
            }
        }
    
    async def _generate_json_report(
        self,
        data: Dict[str, Any],
        manager1_id: int,
        manager2_id: int
    ) -> Path:
        """Generate JSON report."""
        filename = f"h2h_report_{manager1_id}_vs_{manager2_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        file_path = self.reports_dir / filename
        
        with open(file_path, 'w') as f:
            json.dump(data, f, indent=2, default=str)
        
        logger.info(f"Generated JSON report: {file_path}")
        return file_path
    
    async def _generate_csv_report(
        self,
        data: Dict[str, Any],
        manager1_id: int,
        manager2_id: int
    ) -> Path:
        """Generate CSV report with match history."""
        filename = f"h2h_report_{manager1_id}_vs_{manager2_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        file_path = self.reports_dir / filename
        
        with open(file_path, 'w', newline='') as f:
            writer = csv.writer(f)
            
            # Header
            writer.writerow([
                'Gameweek', 'Manager 1', 'Score 1', 'Manager 2', 'Score 2', 
                'Winner', 'Margin'
            ])
            
            # Match data
            for match in data.get('all_matches', []):
                gw = match.get('event', '')
                
                if match.get('entry_1_entry') == manager1_id:
                    m1_name = match.get('entry_1_name', '')
                    m1_score = match.get('entry_1_points', 0)
                    m2_name = match.get('entry_2_name', '')
                    m2_score = match.get('entry_2_points', 0)
                else:
                    m1_name = match.get('entry_2_name', '')
                    m1_score = match.get('entry_2_points', 0)
                    m2_name = match.get('entry_1_name', '')
                    m2_score = match.get('entry_1_points', 0)
                
                if m1_score > m2_score:
                    winner = m1_name
                elif m2_score > m1_score:
                    winner = m2_name
                else:
                    winner = 'Draw'
                
                margin = abs(m1_score - m2_score)
                
                writer.writerow([gw, m1_name, m1_score, m2_name, m2_score, winner, margin])
        
        logger.info(f"Generated CSV report: {file_path}")
        return file_path
    
    async def _generate_pdf_report(
        self,
        data: Dict[str, Any],
        manager1_id: int,
        manager2_id: int
    ) -> Path:
        """Generate PDF report with comprehensive analysis."""
        filename = f"h2h_report_{manager1_id}_vs_{manager2_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        file_path = self.reports_dir / filename
        
        # Create PDF document
        doc = SimpleDocTemplate(str(file_path), pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#38003c'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        manager1_name = data['managers']['manager1']['name']
        manager2_name = data['managers']['manager2']['name']
        
        story.append(Paragraph(f"H2H Season Report", title_style))
        story.append(Paragraph(f"{manager1_name} vs {manager2_name}", styles['Heading2']))
        story.append(Spacer(1, 0.5*inch))
        
        # H2H Record
        record = data['h2h_record']
        record_data = [
            ['H2H Record', 'Wins', 'Draws', 'Losses'],
            [manager1_name, str(record['manager1_wins']), str(record['draws']), str(record['manager2_wins'])],
            [manager2_name, str(record['manager2_wins']), str(record['draws']), str(record['manager1_wins'])]
        ]
        
        record_table = Table(record_data)
        record_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(record_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Season Statistics
        stats = data['statistics']
        if stats:
            story.append(Paragraph("Season Statistics", styles['Heading2']))
            
            stats_data = [
                ['Manager', 'Total Points', 'Avg Points', 'Highest', 'Lowest'],
                [
                    manager1_name,
                    str(stats['manager1']['total_points']),
                    f"{stats['manager1']['avg_points']:.1f}",
                    str(stats['manager1']['highest_score']),
                    str(stats['manager1']['lowest_score'])
                ],
                [
                    manager2_name,
                    str(stats['manager2']['total_points']),
                    f"{stats['manager2']['avg_points']:.1f}",
                    str(stats['manager2']['highest_score']),
                    str(stats['manager2']['lowest_score'])
                ]
            ]
            
            stats_table = Table(stats_data)
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(stats_table)
        
        # Build PDF
        doc.build(story)
        
        logger.info(f"Generated PDF report: {file_path}")
        return file_path
    
    async def generate_league_report(
        self,
        league_id: int,
        output_format: str = "json"
    ) -> Dict[str, Any]:
        """
        Generate a comprehensive league report.
        
        Args:
            league_id: H2H league ID
            output_format: Output format (json, csv, pdf)
            
        Returns:
            Dict containing report data and file path
        """
        logger.info(f"Generating league report for league {league_id}")
        
        # Get league standings
        standings = await self.h2h_analyzer.get_h2h_standings(league_id)
        
        # Get all matches for the season
        all_matches = []
        for gw in range(1, 39):
            try:
                matches = await self.h2h_analyzer.get_h2h_matches(league_id, gw)
                all_matches.extend(matches)
            except Exception as e:
                logger.warning(f"Error fetching matches for GW{gw}: {e}")
        
        report_data = {
            "league_id": league_id,
            "standings": standings,
            "total_matches": len(all_matches),
            "generated_at": datetime.utcnow().isoformat()
        }
        
        # Generate report in requested format
        if output_format.lower() == "json":
            filename = f"league_report_{league_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            file_path = self.reports_dir / filename
            
            with open(file_path, 'w') as f:
                json.dump(report_data, f, indent=2, default=str)
        else:
            raise ValueError(f"League reports currently only support JSON format")
        
        return {
            "status": "success",
            "file_path": str(file_path),
            "format": output_format,
            "data": report_data
        }
    
    async def _generate_excel_report(
        self,
        data: Dict[str, Any],
        manager1_id: int,
        manager2_id: int
    ) -> Path:
        """Generate comprehensive Excel report with multiple sheets and charts."""
        filename = f"h2h_report_{manager1_id}_vs_{manager2_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = self.reports_dir / filename
        
        # Create workbook and worksheets
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        summary_ws = wb.create_sheet("Summary")
        matches_ws = wb.create_sheet("Match History")
        stats_ws = wb.create_sheet("Statistics")
        charts_ws = wb.create_sheet("Charts")
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Summary Sheet
        self._create_excel_summary_sheet(summary_ws, data, header_font, header_fill, border)
        
        # Match History Sheet  
        self._create_excel_matches_sheet(matches_ws, data, header_font, header_fill, border)
        
        # Statistics Sheet
        self._create_excel_stats_sheet(stats_ws, data, header_font, header_fill, border)
        
        # Charts Sheet
        if data.get('all_matches'):
            self._create_excel_charts_sheet(charts_ws, data, wb)
        
        # Save workbook
        wb.save(file_path)
        
        logger.info(f"Generated Excel report: {file_path}")
        return file_path
    
    def _create_excel_summary_sheet(self, ws, data, header_font, header_fill, border):
        """Create summary sheet with key metrics."""
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = "H2H Season Report Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Manager names
        ws['A3'] = "Manager 1:"
        ws['B3'] = data['managers']['manager1']['name']
        ws['A4'] = "Manager 2:"
        ws['B4'] = data['managers']['manager2']['name']
        
        # H2H Record
        ws['A6'] = "H2H Record"
        ws['A6'].font = header_font
        ws['A6'].fill = header_fill
        
        record = data['h2h_record']
        ws['A7'] = "Manager 1 Wins"
        ws['B7'] = record['manager1_wins']
        ws['A8'] = "Draws"
        ws['B8'] = record['draws']
        ws['A9'] = "Manager 2 Wins"
        ws['B9'] = record['manager2_wins']
        
        # Key Statistics
        if data.get('statistics'):
            stats = data['statistics']
            ws['D6'] = "Key Statistics"
            ws['D6'].font = header_font
            ws['D6'].fill = header_fill
            
            ws['D7'] = "Manager 1 Avg Points"
            ws['E7'] = f"{stats['manager1']['avg_points']:.1f}"
            ws['D8'] = "Manager 2 Avg Points"
            ws['E8'] = f"{stats['manager2']['avg_points']:.1f}"
            ws['D9'] = "Avg Margin"
            ws['E9'] = f"{stats['margins']['avg_margin']:.1f}"
        
        # Apply borders to used range
        for row in ws['A1:F15']:
            for cell in row:
                cell.border = border
    
    def _create_excel_matches_sheet(self, ws, data, header_font, header_fill, border):
        """Create match history sheet."""
        # Headers
        headers = ['Gameweek', 'Manager 1', 'Score 1', 'Manager 2', 'Score 2', 'Winner', 'Margin']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Match data
        row = 2
        for match in data.get('all_matches', []):
            gw = match.get('event', '')
            
            if match.get('entry_1_entry') == data['managers']['manager1']['id']:
                m1_name = match.get('entry_1_name', '')
                m1_score = match.get('entry_1_points', 0)
                m2_name = match.get('entry_2_name', '')
                m2_score = match.get('entry_2_points', 0)
            else:
                m1_name = match.get('entry_2_name', '')
                m1_score = match.get('entry_2_points', 0)
                m2_name = match.get('entry_1_name', '')
                m2_score = match.get('entry_1_points', 0)
            
            if m1_score > m2_score:
                winner = m1_name
            elif m2_score > m1_score:
                winner = m2_name
            else:
                winner = 'Draw'
            
            margin = abs(m1_score - m2_score)
            
            ws.cell(row=row, column=1, value=gw)
            ws.cell(row=row, column=2, value=m1_name)
            ws.cell(row=row, column=3, value=m1_score)
            ws.cell(row=row, column=4, value=m2_name)
            ws.cell(row=row, column=5, value=m2_score)
            ws.cell(row=row, column=6, value=winner)
            ws.cell(row=row, column=7, value=margin)
            
            row += 1
        
        # Apply borders and auto-width
        for col in range(1, 8):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].auto_size = True
        
        for row in ws[f'A1:G{row-1}']:
            for cell in row:
                cell.border = border
    
    def _create_excel_stats_sheet(self, ws, data, header_font, header_fill, border):
        """Create detailed statistics sheet."""
        stats = data.get('statistics', {})
        if not stats:
            ws['A1'] = "No statistics available"
            return
        
        # Headers
        ws['A1'] = "Detailed Statistics"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Manager stats table
        headers = ['Metric', data['managers']['manager1']['name'], data['managers']['manager2']['name']]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        metrics = [
            ('Total Points', stats['manager1']['total_points'], stats['manager2']['total_points']),
            ('Average Points', f"{stats['manager1']['avg_points']:.1f}", f"{stats['manager2']['avg_points']:.1f}"),
            ('Highest Score', stats['manager1']['highest_score'], stats['manager2']['highest_score']),
            ('Lowest Score', stats['manager1']['lowest_score'], stats['manager2']['lowest_score'])
        ]
        
        for row, (metric, m1_val, m2_val) in enumerate(metrics, 4):
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=m1_val)
            ws.cell(row=row, column=3, value=m2_val)
        
        # Margin analysis
        if 'margins' in stats:
            ws['A10'] = "Margin Analysis"
            ws['A10'].font = Font(size=12, bold=True)
            
            ws['A11'] = "Average Margin"
            ws['B11'] = f"{stats['margins']['avg_margin']:.1f}"
            ws['A12'] = "Biggest Win Margin"
            ws['B12'] = stats['margins']['biggest_win_margin']
            ws['A13'] = "Biggest Loss Margin"
            ws['B13'] = stats['margins']['biggest_loss_margin']
        
        # Apply borders
        for row in ws['A1:C15']:
            for cell in row:
                cell.border = border
    
    def _create_excel_charts_sheet(self, ws, data, wb):
        """Create charts sheet with visualizations."""
        ws['A1'] = "Data Visualizations"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Points progression chart
        matches = data.get('all_matches', [])
        if matches:
            # Sort matches by gameweek
            sorted_matches = sorted(matches, key=lambda x: x.get('event', 0))
            
            # Create data for line chart
            chart_data = [['Gameweek', data['managers']['manager1']['name'], data['managers']['manager2']['name']]]
            
            for match in sorted_matches[:20]:  # Limit to first 20 matches for clarity
                gw = match.get('event', '')
                if match.get('entry_1_entry') == data['managers']['manager1']['id']:
                    m1_score = match.get('entry_1_points', 0)
                    m2_score = match.get('entry_2_points', 0)
                else:
                    m1_score = match.get('entry_2_points', 0)
                    m2_score = match.get('entry_1_points', 0)
                
                chart_data.append([gw, m1_score, m2_score])
            
            # Write chart data
            for row, data_row in enumerate(chart_data, 3):
                for col, value in enumerate(data_row, 1):
                    ws.cell(row=row, column=col, value=value)
            
            # Create line chart
            chart = LineChart()
            chart.title = "Points Progression"
            chart.style = 13
            chart.x_axis.title = 'Gameweek'
            chart.y_axis.title = 'Points'
            
            data_ref = Reference(ws, min_col=2, min_row=3, max_col=3, max_row=len(chart_data) + 2)
            cats_ref = Reference(ws, min_col=1, min_row=4, max_row=len(chart_data) + 2)
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            
            ws.add_chart(chart, "E3")
    
    async def _generate_html_report(
        self,
        data: Dict[str, Any],
        manager1_id: int,
        manager2_id: int
    ) -> Path:
        """Generate interactive HTML report with Plotly charts."""
        filename = f"h2h_report_{manager1_id}_vs_{manager2_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
        file_path = self.reports_dir / filename
        
        # Create interactive visualizations
        figures = []
        
        # Points progression chart
        if data.get('all_matches'):
            matches = sorted(data['all_matches'], key=lambda x: x.get('event', 0))
            
            gameweeks = []
            m1_scores = []
            m2_scores = []
            
            for match in matches:
                gw = match.get('event', 0)
                if match.get('entry_1_entry') == manager1_id:
                    m1_score = match.get('entry_1_points', 0)
                    m2_score = match.get('entry_2_points', 0)
                else:
                    m1_score = match.get('entry_2_points', 0)
                    m2_score = match.get('entry_1_points', 0)
                
                gameweeks.append(f"GW{gw}")
                m1_scores.append(m1_score)
                m2_scores.append(m2_score)
            
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=gameweeks, 
                y=m1_scores,
                mode='lines+markers',
                name=data['managers']['manager1']['name'],
                line=dict(color='#1f77b4', width=3)
            ))
            fig.add_trace(go.Scatter(
                x=gameweeks, 
                y=m2_scores,
                mode='lines+markers',
                name=data['managers']['manager2']['name'],
                line=dict(color='#ff7f0e', width=3)
            ))
            
            fig.update_layout(
                title="Points Progression Throughout Season",
                xaxis_title="Gameweek",
                yaxis_title="Points",
                template="plotly_white",
                height=500
            )
            
            figures.append(fig)
        
        # Win/Loss distribution
        record = data.get('h2h_record', {})
        if record:
            labels = [
                f"{data['managers']['manager1']['name']} Wins",
                'Draws',
                f"{data['managers']['manager2']['name']} Wins"
            ]
            values = [record.get('manager1_wins', 0), record.get('draws', 0), record.get('manager2_wins', 0)]
            
            fig = go.Figure(data=[go.Pie(labels=labels, values=values, hole=.3)])
            fig.update_layout(
                title="H2H Win Distribution",
                template="plotly_white",
                height=400
            )
            
            figures.append(fig)
        
        # Generate HTML content
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>H2H Season Report - {data['managers']['manager1']['name']} vs {data['managers']['manager2']['name']}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                .header {{ text-align: center; margin-bottom: 30px; }}
                .summary {{ background: #f8f9fa; padding: 20px; border-radius: 8px; margin-bottom: 30px; }}
                .stat-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; }}
                .stat-card {{ background: white; padding: 15px; border-radius: 8px; border: 1px solid #dee2e6; }}
                .chart-container {{ margin: 30px 0; }}
            </style>
            <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
        </head>
        <body>
            <div class="header">
                <h1>H2H Season Report</h1>
                <h2>{data['managers']['manager1']['name']} vs {data['managers']['manager2']['name']}</h2>
                <p>Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            </div>
            
            <div class="summary">
                <h3>Season Summary</h3>
                <div class="stat-grid">
                    <div class="stat-card">
                        <h4>H2H Record</h4>
                        <p>{record.get('manager1_wins', 0)} - {record.get('draws', 0)} - {record.get('manager2_wins', 0)}</p>
                    </div>
        """
        
        # Add statistics if available
        if data.get('statistics'):
            stats = data['statistics']
            html_content += f"""
                    <div class="stat-card">
                        <h4>{data['managers']['manager1']['name']} Stats</h4>
                        <p>Avg: {stats['manager1']['avg_points']:.1f} pts</p>
                        <p>High: {stats['manager1']['highest_score']} pts</p>
                    </div>
                    <div class="stat-card">
                        <h4>{data['managers']['manager2']['name']} Stats</h4>
                        <p>Avg: {stats['manager2']['avg_points']:.1f} pts</p>
                        <p>High: {stats['manager2']['highest_score']} pts</p>
                    </div>
            """
        
        html_content += """
                </div>
            </div>
        """
        
        # Add charts
        for i, fig in enumerate(figures):
            chart_html = plot(fig, output_type='div', include_plotlyjs=False)
            html_content += f'<div class="chart-container">{chart_html}</div>'
        
        html_content += """
        </body>
        </html>
        """
        
        # Write HTML file
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        logger.info(f"Generated HTML report: {file_path}")
        return file_path
    
    async def _generate_enhanced_pdf_report(
        self,
        data: Dict[str, Any],
        manager1_id: int,
        manager2_id: int
    ) -> Path:
        """Generate enhanced PDF report with better formatting and charts."""
        filename = f"h2h_report_{manager1_id}_vs_{manager2_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        file_path = self.reports_dir / filename
        
        # Create PDF document
        doc = SimpleDocTemplate(str(file_path), pagesize=A4, topMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=28,
            textColor=colors.HexColor('#38003c'),
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=18,
            textColor=colors.HexColor('#00ff85'),
            spaceAfter=15,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        section_style = ParagraphStyle(
            'SectionHeader',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#38003c'),
            spaceBefore=20,
            spaceAfter=10,
            fontName='Helvetica-Bold'
        )
        
        # Title and subtitle
        manager1_name = data['managers']['manager1']['name']
        manager2_name = data['managers']['manager2']['name']
        
        story.append(Paragraph("FPL H2H SEASON ANALYSIS", title_style))
        story.append(Paragraph(f"{manager1_name} vs {manager2_name}", subtitle_style))
        story.append(Spacer(1, 0.3*inch))
        
        # Executive Summary
        story.append(Paragraph("Executive Summary", section_style))
        
        record = data['h2h_record']
        total_matches = record['manager1_wins'] + record['draws'] + record['manager2_wins']
        
        if record['manager1_wins'] > record['manager2_wins']:
            dominant_manager = manager1_name
            win_percentage = (record['manager1_wins'] / total_matches * 100) if total_matches > 0 else 0
        elif record['manager2_wins'] > record['manager1_wins']:
            dominant_manager = manager2_name
            win_percentage = (record['manager2_wins'] / total_matches * 100) if total_matches > 0 else 0
        else:
            dominant_manager = "Neither (tied)"
            win_percentage = 50
        
        summary_text = f"""
        This comprehensive analysis covers {total_matches} head-to-head matches between {manager1_name} and {manager2_name}.
        
        <b>Key Findings:</b><br/>
        • Dominant Manager: {dominant_manager} ({win_percentage:.1f}% win rate)<br/>
        • Total Matches Played: {total_matches}<br/>
        • Competitive Balance: {'High' if abs(record['manager1_wins'] - record['manager2_wins']) <= 2 else 'Low'}<br/>
        """
        
        if data.get('statistics'):
            stats = data['statistics']
            avg_diff = abs(stats['manager1']['avg_points'] - stats['manager2']['avg_points'])
            summary_text += f"• Average Points Difference: {avg_diff:.1f} points per match<br/>"
        
        story.append(Paragraph(summary_text, styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Enhanced H2H Record Table
        story.append(Paragraph("Head-to-Head Record", section_style))
        
        record_data = [
            ['', 'Wins', 'Draws', 'Losses', 'Win %'],
            [
                manager1_name, 
                str(record['manager1_wins']), 
                str(record['draws']), 
                str(record['manager2_wins']),
                f"{(record['manager1_wins'] / total_matches * 100):.1f}%" if total_matches > 0 else "0%"
            ],
            [
                manager2_name, 
                str(record['manager2_wins']), 
                str(record['draws']), 
                str(record['manager1_wins']),
                f"{(record['manager2_wins'] / total_matches * 100):.1f}%" if total_matches > 0 else "0%"
            ]
        ]
        
        record_table = Table(record_data, colWidths=[2*inch, 1*inch, 1*inch, 1*inch, 1*inch])
        record_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#38003c')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#e8f5e8')),
            ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#fff5e6')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10)
        ]))
        
        story.append(record_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Enhanced Season Statistics
        if data.get('statistics'):
            stats = data['statistics']
            story.append(Paragraph("Detailed Season Statistics", section_style))
            
            stats_data = [
                ['Metric', manager1_name, manager2_name, 'Difference'],
                [
                    'Total Points',
                    f"{stats['manager1']['total_points']:,}",
                    f"{stats['manager2']['total_points']:,}",
                    f"{abs(stats['manager1']['total_points'] - stats['manager2']['total_points']):,}"
                ],
                [
                    'Average Points',
                    f"{stats['manager1']['avg_points']:.1f}",
                    f"{stats['manager2']['avg_points']:.1f}",
                    f"{abs(stats['manager1']['avg_points'] - stats['manager2']['avg_points']):.1f}"
                ],
                [
                    'Highest Score',
                    str(stats['manager1']['highest_score']),
                    str(stats['manager2']['highest_score']),
                    str(abs(stats['manager1']['highest_score'] - stats['manager2']['highest_score']))
                ],
                [
                    'Lowest Score',
                    str(stats['manager1']['lowest_score']),
                    str(stats['manager2']['lowest_score']),
                    str(abs(stats['manager1']['lowest_score'] - stats['manager2']['lowest_score']))
                ]
            ]
            
            stats_table = Table(stats_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#38003c')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9)
            ]))
            
            story.append(stats_table)
            story.append(Spacer(1, 0.3*inch))
        
        # Key Battles Section
        if data.get('key_battles'):
            battles = data['key_battles']
            story.append(Paragraph("Memorable Matches", section_style))
            
            battle_text = ""
            if battles.get('biggest_manager1_win'):
                match = battles['biggest_manager1_win']
                gw = match.get('event', 'Unknown')
                if match.get('entry_1_entry') == manager1_id:
                    score = f"{match.get('entry_1_points', 0)} - {match.get('entry_2_points', 0)}"
                else:
                    score = f"{match.get('entry_2_points', 0)} - {match.get('entry_1_points', 0)}"
                battle_text += f"<b>Biggest {manager1_name} Victory:</b> GW{gw} ({score})<br/>"
            
            if battles.get('biggest_manager2_win'):
                match = battles['biggest_manager2_win']
                gw = match.get('event', 'Unknown')
                if match.get('entry_1_entry') == manager2_id:
                    score = f"{match.get('entry_1_points', 0)} - {match.get('entry_2_points', 0)}"
                else:
                    score = f"{match.get('entry_2_points', 0)} - {match.get('entry_1_points', 0)}"
                battle_text += f"<b>Biggest {manager2_name} Victory:</b> GW{gw} ({score})<br/>"
            
            if battles.get('closest_match'):
                match = battles['closest_match']
                gw = match.get('event', 'Unknown')
                if match.get('entry_1_entry') == manager1_id:
                    score = f"{match.get('entry_1_points', 0)} - {match.get('entry_2_points', 0)}"
                else:
                    score = f"{match.get('entry_2_points', 0)} - {match.get('entry_1_points', 0)}"
                battle_text += f"<b>Closest Match:</b> GW{gw} ({score})<br/>"
            
            if battle_text:
                story.append(Paragraph(battle_text, styles['Normal']))
                story.append(Spacer(1, 0.2*inch))
        
        # Form Analysis
        if data.get('form_analysis'):
            form = data['form_analysis']
            story.append(Paragraph("Form Analysis", section_style))
            
            form_text = ""
            if form.get('best_winning_streak'):
                streaks = form['best_winning_streak']
                form_text += f"<b>Best Winning Streaks:</b><br/>"
                form_text += f"• {manager1_name}: {streaks.get('manager1', 0)} matches<br/>"
                form_text += f"• {manager2_name}: {streaks.get('manager2', 0)} matches<br/>"
            
            if form.get('current_form'):
                current = form['current_form']
                form_text += f"<b>Recent Form (Last 5 matches):</b><br/>"
                form_text += f"• {manager1_name}: {current.get('manager1', 'N/A')}<br/>"
                form_text += f"• {manager2_name}: {current.get('manager2', 'N/A')}<br/>"
            
            if form_text:
                story.append(Paragraph(form_text, styles['Normal']))
        
        # Footer
        story.append(Spacer(1, 0.5*inch))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        story.append(Paragraph(f"Generated on {datetime.now().strftime('%Y-%m-%d at %H:%M:%S')} | FPL H2H Analyzer", footer_style))
        
        # Build PDF
        doc.build(story)
        
        logger.info(f"Generated enhanced PDF report: {file_path}")
        return file_path